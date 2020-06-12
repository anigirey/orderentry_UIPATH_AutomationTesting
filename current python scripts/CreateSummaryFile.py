"""
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import xlrd
import os
from os import path


def main():

    currDir = os.getcwd()
    loc = currDir+("/Output/outputvars.xlsx")
    wBook = xlrd.open_workbook(loc)
    wSheet = wBook.sheet_by_index(0)
    testEnv = str(wSheet.cell_value(0,3))
    testTim = str(wSheet.cell_value(0,1))
    testDat = str(wSheet.cell_value(0,0))
    matchedSheet = int()

    if not path.exists(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx'):
        summaryWorkbook = Workbook()
        summaryWorksheet = summaryWorkbook.create_sheet(testTim)
        dust = summaryWorkbook.get_sheet_by_name('Sheet')
        summaryWorkbook.remove_sheet(dust)

        celBold = Font(size=13,bold=True,underline='single')
        celBoldL = Font(size=11,bold=True)
        celCent = Alignment(horizontal='center')
        celLeft = Alignment(horizontal='left')
        
        summaryWorksheet.column_dimensions['A'].width=47
        summaryWorksheet['A1'].font=celBold
        summaryWorksheet['A1'].alignment=celCent
        summaryWorksheet['A1']='Step Detail'
        
        summaryWorksheet['A2'].font=celBoldL
        summaryWorksheet['A2'].alignment=celLeft
        summaryWorksheet['A2']='Sign On Page opened'

        summaryWorksheet['A3'].font=celBoldL
        summaryWorksheet['A3'].alignment=celLeft
        summaryWorksheet['A3']='Home Page opened'

        summaryWorksheet['A4'].font=celBoldL
        summaryWorksheet['A4'].alignment=celLeft
        summaryWorksheet['A4']='Customer Information Page opened'

        summaryWorksheet['A5'].font=celBoldL
        summaryWorksheet['A5'].alignment=celLeft
        summaryWorksheet['A5']='Win Back/Win Over Page opened'

        summaryWorksheet['A6'].font=celBoldL
        summaryWorksheet['A6'].alignment=celLeft
        summaryWorksheet['A6']='Service Address Validation Page opened'

        summaryWorksheet['A7'].font=celBoldL
        summaryWorksheet['A7'].alignment=celLeft
        summaryWorksheet['A7']='Facility Check and Results Page opened'

        summaryWorksheet['A8'].font=celBoldL
        summaryWorksheet['A8'].alignment=celLeft
        summaryWorksheet['A8']='Primary Listing Page opened'

        summaryWorksheet['A9'].font=celBoldL
        summaryWorksheet['A9'].alignment=celLeft
        summaryWorksheet['A9']='Product Pricing Page opened'

        summaryWorksheet['A10'].font=celBoldL
        summaryWorksheet['A10'].alignment=celLeft
        summaryWorksheet['A10']='Billing Information Page opened'

        summaryWorksheet['A11'].font=celBoldL
        summaryWorksheet['A11'].alignment=celLeft
        summaryWorksheet['A11']='Business Credit Application Page opened'

        summaryWorksheet['A12'].font=celBoldL
        summaryWorksheet['A12'].alignment=celLeft
        summaryWorksheet['A12']='Credit Information Page opened'

        summaryWorksheet['A13'].font=celBoldL
        summaryWorksheet['A13'].alignment=celLeft
        summaryWorksheet['A13']='Credit Decision Page opened'

        summaryWorksheet['A14'].font=celBoldL
        summaryWorksheet['A14'].alignment=celLeft
        summaryWorksheet['A14']='Service and Equipment Page opened'

        summaryWorksheet['A15'].font=celBoldL
        summaryWorksheet['A15'].alignment=celLeft
        summaryWorksheet['A15']='Internet Ordering Page opened'

        summaryWorksheet['A16'].font=celBoldL
        summaryWorksheet['A16'].alignment=celLeft
        summaryWorksheet['A16']='Product Summary Page opened'

        summaryWorksheet['A17'].font=celBoldL
        summaryWorksheet['A17'].alignment=celLeft
        summaryWorksheet['A17']='Configure Product Page opened'

        summaryWorksheet['A18'].font=celBoldL
        summaryWorksheet['A18'].alignment=celLeft
        summaryWorksheet['A18']='Configure Order Page opened'

        summaryWorksheet['A19'].font=celBoldL
        summaryWorksheet['A19'].alignment=celLeft
        summaryWorksheet['A19']='Configure OLFIDs Page opened'

        summaryWorksheet['A20'].font=celBoldL
        summaryWorksheet['A20'].alignment=celLeft
        summaryWorksheet['A20']='Appointment Scheduler Page opened'

        summaryWorksheet['A21'].font=celBoldL
        summaryWorksheet['A21'].alignment=celLeft
        summaryWorksheet['A21']='Deposit/Advance Payment Page opened'

        summaryWorksheet['A22'].font=celBoldL
        summaryWorksheet['A22'].alignment=celLeft
        summaryWorksheet['A22']='Deposit/Advance Payment Information Page opened'

        summaryWorksheet['A23'].font=celBoldL
        summaryWorksheet['A23'].alignment=celLeft
        summaryWorksheet['A23']='Deposit/Advance Payment Success Page opened'

        summaryWorksheet['A24'].font=celBoldL
        summaryWorksheet['A24'].alignment=celLeft
        summaryWorksheet['A24']='Order Detail Page opened'

        summaryWorksheet['A25'].font=celBoldL
        summaryWorksheet['A25'].alignment=celLeft
        summaryWorksheet['A25']='Order Validation Confirmation'

        summaryWorksheet['A27'].font=celBoldL
        summaryWorksheet['A27'].alignment=celLeft
        summaryWorksheet['A27']='Reason for Failure'

        # 1FR init
        summaryWorksheet.column_dimensions['B'].width=15
        summaryWorksheet['B1'].font=celBold
        summaryWorksheet['B1'].alignment=celCent
        summaryWorksheet['B1']='1FR'
        summaryWorksheet['B2']='---'
        summaryWorksheet['B3']='---'
        summaryWorksheet['B4']='---'
        summaryWorksheet['B5']='---'
        summaryWorksheet['B6']='---'
        summaryWorksheet['B7']='---'
        summaryWorksheet['B8']='---'
        summaryWorksheet['B9']='---'
        summaryWorksheet['B10']='---'
        summaryWorksheet['B11']='---'
        summaryWorksheet['B12']='---'
        summaryWorksheet['B13']='---'
        summaryWorksheet['B14']='---'
        summaryWorksheet['B15']='---'
        summaryWorksheet['B16']='---'
        summaryWorksheet['B17']='---'
        summaryWorksheet['B18']='---'
        summaryWorksheet['B19']='---'
        summaryWorksheet['B20']='---'
        summaryWorksheet['B21']='---'
        summaryWorksheet['B22']='---'
        summaryWorksheet['B23']='---'
        summaryWorksheet['B24']='---'
        summaryWorksheet['B25']='---'
        summaryWorksheet['B27']='---'
        
        # 1FB init
        summaryWorksheet.column_dimensions['C'].width=15
        summaryWorksheet['C1'].font=celBold
        summaryWorksheet['C1'].alignment=celCent
        summaryWorksheet['C1']='1FB'
        summaryWorksheet['C2']='---'
        summaryWorksheet['C3']='---'
        summaryWorksheet['C4']='---'
        summaryWorksheet['C5']='---'
        summaryWorksheet['C6']='---'
        summaryWorksheet['C7']='---'
        summaryWorksheet['C8']='---'
        summaryWorksheet['C9']='---'
        summaryWorksheet['C10']='---'
        summaryWorksheet['C11']='---'
        summaryWorksheet['C12']='---'
        summaryWorksheet['C13']='---'
        summaryWorksheet['C14']='---'
        summaryWorksheet['C15']='---'
        summaryWorksheet['C16']='---'
        summaryWorksheet['C17']='---'
        summaryWorksheet['C18']='---'
        summaryWorksheet['C19']='---'
        summaryWorksheet['C20']='---'
        summaryWorksheet['C21']='---'
        summaryWorksheet['C22']='---'
        summaryWorksheet['C23']='---'
        summaryWorksheet['C24']='---'
        summaryWorksheet['C25']='---'
        summaryWorksheet['C27']='---'

        # HSI/GFR init
        summaryWorksheet.column_dimensions['D'].width=15
        summaryWorksheet['D1'].font=celBold
        summaryWorksheet['D1'].alignment=celCent
        summaryWorksheet['D1']='GFR'
        summaryWorksheet['D2']='---'
        summaryWorksheet['D3']='---'
        summaryWorksheet['D4']='---'
        summaryWorksheet['D5']='---'
        summaryWorksheet['D6']='---'
        summaryWorksheet['D7']='---'
        summaryWorksheet['D8']='---'
        summaryWorksheet['D9']='---'
        summaryWorksheet['D10']='---'
        summaryWorksheet['D11']='---'
        summaryWorksheet['D12']='---'
        summaryWorksheet['D13']='---'
        summaryWorksheet['D14']='---'
        summaryWorksheet['D15']='---'
        summaryWorksheet['D16']='---'
        summaryWorksheet['D17']='---'
        summaryWorksheet['D18']='---'
        summaryWorksheet['D19']='---'
        summaryWorksheet['D20']='---'
        summaryWorksheet['D21']='---'
        summaryWorksheet['D22']='---'
        summaryWorksheet['D23']='---'
        summaryWorksheet['D24']='---'
        summaryWorksheet['D25']='---'
        summaryWorksheet['D27']='---'

        # GFB init
        summaryWorksheet.column_dimensions['E'].width=15
        summaryWorksheet['E1'].font=celBold
        summaryWorksheet['E1'].alignment=celCent
        summaryWorksheet['E1']='GFB'
        summaryWorksheet['E2']='---'
        summaryWorksheet['E3']='---'
        summaryWorksheet['E4']='---'
        summaryWorksheet['E5']='---'
        summaryWorksheet['E6']='---'
        summaryWorksheet['E7']='---'
        summaryWorksheet['E8']='---'
        summaryWorksheet['E9']='---'
        summaryWorksheet['E10']='---'
        summaryWorksheet['E11']='---'
        summaryWorksheet['E12']='---'
        summaryWorksheet['E13']='---'
        summaryWorksheet['E14']='---'
        summaryWorksheet['E15']='---'
        summaryWorksheet['E16']='---'
        summaryWorksheet['E17']='---'
        summaryWorksheet['E18']='---'
        summaryWorksheet['E19']='---'
        summaryWorksheet['E20']='---'
        summaryWorksheet['E21']='---'
        summaryWorksheet['E22']='---'
        summaryWorksheet['E23']='---'
        summaryWorksheet['E24']='---'
        summaryWorksheet['E25']='---'
        summaryWorksheet['E27']='---'

        # PRISM init
        summaryWorksheet.column_dimensions['F'].width=15
        summaryWorksheet['F1'].font=celBold
        summaryWorksheet['F1'].alignment=celCent
        summaryWorksheet['F1']='PRISM'
        summaryWorksheet['F2']='---'
        summaryWorksheet['F3']='---'
        summaryWorksheet['F4']='---'
        summaryWorksheet['F5']='---'
        summaryWorksheet['F6']='---'
        summaryWorksheet['F7']='---'
        summaryWorksheet['F8']='---'
        summaryWorksheet['F9']='---'
        summaryWorksheet['F10']='---'
        summaryWorksheet['F11']='---'
        summaryWorksheet['F12']='---'
        summaryWorksheet['F13']='---'
        summaryWorksheet['F14']='---'
        summaryWorksheet['F15']='---'
        summaryWorksheet['F16']='---'
        summaryWorksheet['F17']='---'
        summaryWorksheet['F18']='---'
        summaryWorksheet['F19']='---'
        summaryWorksheet['F20']='---'
        summaryWorksheet['F21']='---'
        summaryWorksheet['F22']='---'
        summaryWorksheet['F23']='---'
        summaryWorksheet['F24']='---'
        summaryWorksheet['F25']='---'
        summaryWorksheet['F27']='---'

        summaryWorkbook.save(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx')
        summaryWorkbook.close()

    else:
        
        summaryWorkbook = openpyxl.load_workbook(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx')
                
        for i in summaryWorkbook.sheetnames:
            if (str(i) == testTim):
                matchedSheet = int(1)
                break
            else:
                matchedSheet = int(0)

        if (matchedSheet == int(0)):

            summaryWorksheet = summaryWorkbook.create_sheet(testTim)
            
            celBold = Font(size=13,bold=True,underline='single')
            celBoldL = Font(size=11,bold=True)
            celCent = Alignment(horizontal='center')
            celLeft = Alignment(horizontal='left')
            
            summaryWorksheet.column_dimensions['A'].width=47
            summaryWorksheet['A1'].font=celBold
            summaryWorksheet['A1'].alignment=celCent
            summaryWorksheet['A1']='Step Detail'
            
            summaryWorksheet['A2'].font=celBoldL
            summaryWorksheet['A2'].alignment=celLeft
            summaryWorksheet['A2']='Sign On Page opened'

            summaryWorksheet['A3'].font=celBoldL
            summaryWorksheet['A3'].alignment=celLeft
            summaryWorksheet['A3']='Home Page opened'

            summaryWorksheet['A4'].font=celBoldL
            summaryWorksheet['A4'].alignment=celLeft
            summaryWorksheet['A4']='Customer Information Page opened'

            summaryWorksheet['A5'].font=celBoldL
            summaryWorksheet['A5'].alignment=celLeft
            summaryWorksheet['A5']='Win Back/Win Over Page opened'

            summaryWorksheet['A6'].font=celBoldL
            summaryWorksheet['A6'].alignment=celLeft
            summaryWorksheet['A6']='Service Address Validation Page opened'

            summaryWorksheet['A7'].font=celBoldL
            summaryWorksheet['A7'].alignment=celLeft
            summaryWorksheet['A7']='Facility Check and Results Page opened'

            summaryWorksheet['A8'].font=celBoldL
            summaryWorksheet['A8'].alignment=celLeft
            summaryWorksheet['A8']='Primary Listing Page opened'

            summaryWorksheet['A9'].font=celBoldL
            summaryWorksheet['A9'].alignment=celLeft
            summaryWorksheet['A9']='Product Pricing Page opened'

            summaryWorksheet['A10'].font=celBoldL
            summaryWorksheet['A10'].alignment=celLeft
            summaryWorksheet['A10']='Billing Information Page opened'

            summaryWorksheet['A11'].font=celBoldL
            summaryWorksheet['A11'].alignment=celLeft
            summaryWorksheet['A11']='Business Credit Application Page opened'

            summaryWorksheet['A12'].font=celBoldL
            summaryWorksheet['A12'].alignment=celLeft
            summaryWorksheet['A12']='Credit Information Page opened'

            summaryWorksheet['A13'].font=celBoldL
            summaryWorksheet['A13'].alignment=celLeft
            summaryWorksheet['A13']='Credit Decision Page opened'

            summaryWorksheet['A14'].font=celBoldL
            summaryWorksheet['A14'].alignment=celLeft
            summaryWorksheet['A14']='Service and Equipment Page opened'

            summaryWorksheet['A15'].font=celBoldL
            summaryWorksheet['A15'].alignment=celLeft
            summaryWorksheet['A15']='Internet Ordering Page opened'

            summaryWorksheet['A16'].font=celBoldL
            summaryWorksheet['A16'].alignment=celLeft
            summaryWorksheet['A16']='Product Summary Page opened'

            summaryWorksheet['A17'].font=celBoldL
            summaryWorksheet['A17'].alignment=celLeft
            summaryWorksheet['A17']='Configure Product Page opened'

            summaryWorksheet['A18'].font=celBoldL
            summaryWorksheet['A18'].alignment=celLeft
            summaryWorksheet['A18']='Configure Order Page opened'

            summaryWorksheet['A19'].font=celBoldL
            summaryWorksheet['A19'].alignment=celLeft
            summaryWorksheet['A19']='Configure OLFIDs Page opened'

            summaryWorksheet['A20'].font=celBoldL
            summaryWorksheet['A20'].alignment=celLeft
            summaryWorksheet['A20']='Appointment Scheduler Page opened'

            summaryWorksheet['A21'].font=celBoldL
            summaryWorksheet['A21'].alignment=celLeft
            summaryWorksheet['A21']='Deposit/Advance Payment Page opened'

            summaryWorksheet['A22'].font=celBoldL
            summaryWorksheet['A22'].alignment=celLeft
            summaryWorksheet['A22']='Deposit/Advance Payment Information Page opened'

            summaryWorksheet['A23'].font=celBoldL
            summaryWorksheet['A23'].alignment=celLeft
            summaryWorksheet['A23']='Deposit/Advance Payment Success Page opened'

            summaryWorksheet['A24'].font=celBoldL
            summaryWorksheet['A24'].alignment=celLeft
            summaryWorksheet['A24']='Order Detail Page opened'

            summaryWorksheet['A25'].font=celBoldL
            summaryWorksheet['A25'].alignment=celLeft
            summaryWorksheet['A25']='Order Validation Confirmation'

            summaryWorksheet['A27'].font=celBoldL
            summaryWorksheet['A27'].alignment=celLeft
            summaryWorksheet['A27']='Reason for Failure'

            # 1FR init
            summaryWorksheet.column_dimensions['B'].width=15
            summaryWorksheet['B1'].font=celBold
            summaryWorksheet['B1'].alignment=celCent
            summaryWorksheet['B1']='1FR'
            summaryWorksheet['B2']='---'
            summaryWorksheet['B3']='---'
            summaryWorksheet['B4']='---'
            summaryWorksheet['B5']='---'
            summaryWorksheet['B6']='---'
            summaryWorksheet['B7']='---'
            summaryWorksheet['B8']='---'
            summaryWorksheet['B9']='---'
            summaryWorksheet['B10']='---'
            summaryWorksheet['B11']='---'
            summaryWorksheet['B12']='---'
            summaryWorksheet['B13']='---'
            summaryWorksheet['B14']='---'
            summaryWorksheet['B15']='---'
            summaryWorksheet['B16']='---'
            summaryWorksheet['B17']='---'
            summaryWorksheet['B18']='---'
            summaryWorksheet['B19']='---'
            summaryWorksheet['B20']='---'
            summaryWorksheet['B21']='---'
            summaryWorksheet['B22']='---'
            summaryWorksheet['B23']='---'
            summaryWorksheet['B24']='---'
            summaryWorksheet['B25']='---'
            summaryWorksheet['B27']='---'

            # 1FB init
            summaryWorksheet.column_dimensions['C'].width=15
            summaryWorksheet['C1'].font=celBold
            summaryWorksheet['C1'].alignment=celCent
            summaryWorksheet['C1']='1FB'
            summaryWorksheet['C2']='---'
            summaryWorksheet['C3']='---'
            summaryWorksheet['C4']='---'
            summaryWorksheet['C5']='---'
            summaryWorksheet['C6']='---'
            summaryWorksheet['C7']='---'
            summaryWorksheet['C8']='---'
            summaryWorksheet['C9']='---'
            summaryWorksheet['C10']='---'
            summaryWorksheet['C11']='---'
            summaryWorksheet['C12']='---'
            summaryWorksheet['C13']='---'
            summaryWorksheet['C14']='---'
            summaryWorksheet['C15']='---'
            summaryWorksheet['C16']='---'
            summaryWorksheet['C17']='---'
            summaryWorksheet['C18']='---'
            summaryWorksheet['C19']='---'
            summaryWorksheet['C20']='---'
            summaryWorksheet['C21']='---'
            summaryWorksheet['C22']='---'
            summaryWorksheet['C23']='---'
            summaryWorksheet['C24']='---'
            summaryWorksheet['C25']='---'
            summaryWorksheet['C27']='---'

            # HSI init
            summaryWorksheet.column_dimensions['D'].width=15
            summaryWorksheet['D1'].font=celBold
            summaryWorksheet['D1'].alignment=celCent
            summaryWorksheet['D1']='GFR'
            summaryWorksheet['D2']='---'
            summaryWorksheet['D3']='---'
            summaryWorksheet['D4']='---'
            summaryWorksheet['D5']='---'
            summaryWorksheet['D6']='---'
            summaryWorksheet['D7']='---'
            summaryWorksheet['D8']='---'
            summaryWorksheet['D9']='---'
            summaryWorksheet['D10']='---'
            summaryWorksheet['D11']='---'
            summaryWorksheet['D12']='---'
            summaryWorksheet['D13']='---'
            summaryWorksheet['D14']='---'
            summaryWorksheet['D15']='---'
            summaryWorksheet['D16']='---'
            summaryWorksheet['D17']='---'
            summaryWorksheet['D18']='---'
            summaryWorksheet['D19']='---'
            summaryWorksheet['D20']='---'
            summaryWorksheet['D21']='---'
            summaryWorksheet['D22']='---'
            summaryWorksheet['D23']='---'
            summaryWorksheet['D24']='---'
            summaryWorksheet['D25']='---'
            summaryWorksheet['D27']='---'

            # GFB init
            summaryWorksheet.column_dimensions['E'].width=15
            summaryWorksheet['E1'].font=celBold
            summaryWorksheet['E1'].alignment=celCent
            summaryWorksheet['E1']='GFB'
            summaryWorksheet['E2']='---'
            summaryWorksheet['E3']='---'
            summaryWorksheet['E4']='---'
            summaryWorksheet['E5']='---'
            summaryWorksheet['E6']='---'
            summaryWorksheet['E7']='---'
            summaryWorksheet['E8']='---'
            summaryWorksheet['E9']='---'
            summaryWorksheet['E10']='---'
            summaryWorksheet['E11']='---'
            summaryWorksheet['E12']='---'
            summaryWorksheet['E13']='---'
            summaryWorksheet['E14']='---'
            summaryWorksheet['E15']='---'
            summaryWorksheet['E16']='---'
            summaryWorksheet['E17']='---'
            summaryWorksheet['E18']='---'
            summaryWorksheet['E19']='---'
            summaryWorksheet['E20']='---'
            summaryWorksheet['E21']='---'
            summaryWorksheet['E22']='---'
            summaryWorksheet['E23']='---'
            summaryWorksheet['E24']='---'
            summaryWorksheet['E25']='---'
            summaryWorksheet['E27']='---'

            # PRISM init
            summaryWorksheet.column_dimensions['F'].width=15
            summaryWorksheet['F1'].font=celBold
            summaryWorksheet['F1'].alignment=celCent
            summaryWorksheet['F1']='PRISM'
            summaryWorksheet['F2']='---'
            summaryWorksheet['F3']='---'
            summaryWorksheet['F4']='---'
            summaryWorksheet['F5']='---'
            summaryWorksheet['F6']='---'
            summaryWorksheet['F7']='---'
            summaryWorksheet['F8']='---'
            summaryWorksheet['F9']='---'
            summaryWorksheet['F10']='---'
            summaryWorksheet['F11']='---'
            summaryWorksheet['F12']='---'
            summaryWorksheet['F13']='---'
            summaryWorksheet['F14']='---'
            summaryWorksheet['F15']='---'
            summaryWorksheet['F16']='---'
            summaryWorksheet['F17']='---'
            summaryWorksheet['F18']='---'
            summaryWorksheet['F19']='---'
            summaryWorksheet['F20']='---'
            summaryWorksheet['F21']='---'
            summaryWorksheet['F22']='---'
            summaryWorksheet['F23']='---'
            summaryWorksheet['F24']='---'
            summaryWorksheet['F25']='---'
            summaryWorksheet['F27']='---'

            summaryWorkbook.save(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx')
            summaryWorkbook.close()


main()
