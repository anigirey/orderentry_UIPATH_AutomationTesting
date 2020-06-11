"""

"""


import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.workbook import Workbook
import xlrd
import os
from os import path


def main():

    currDir = os.getcwd()
    loc = (str(currDir) + ("/Output/outputvars.xlsx"))
    wBook = xlrd.open_workbook(loc)
    wSheet = wBook.sheet_by_index(0)
    ordType = str(wSheet.cell_value(0,4))
    testEnv = str(wSheet.cell_value(0,3))
    testTim = str(wSheet.cell_value(0,1))
    testDat = str(wSheet.cell_value(0,0))
    matchedSheet = int()

    sanityWorkbook = openpyxl.load_workbook(currDir+"/Results/"+testDat+"/"+
                                            testEnv+'/'+testEnv+"_SR.xlsx")
    for i in sanityWorkbook.sheetnames:
        if (str(i) == testTim):
            matchedSheet = int(1)
            break
        else:
            matchedSheet = int(0)

    if (matchedSheet == int(0)):
        sanityWorksheet = sanityWorkbook.create_sheet(testTim)

        celBold = Font(size=11,bold=True,underline='single')
        celCent = Alignment(horizontal='center')
        
        sanityWorksheet.column_dimensions['A'].width = 14
        sanityWorksheet['A1'].font = celBold
        sanityWorksheet['A1'].alignment = celCent
        sanityWorksheet['A1'] = 'Order Type'

        sanityWorksheet.column_dimensions['B'].width = 14
        sanityWorksheet['B1'].font = celBold
        sanityWorksheet['B1'].alignment = celCent
        sanityWorksheet['B1'] = 'Order Num'

        sanityWorksheet.column_dimensions['C'].width = 14
        sanityWorksheet['C1'].font = celBold
        sanityWorksheet['C1'].alignment = celCent
        sanityWorksheet['C1'] = 'TN'

        sanityWorksheet.column_dimensions['D'].width = 14
        sanityWorksheet['D1'].font = celBold
        sanityWorksheet['D1'].alignment = celCent
        sanityWorksheet['D1'] = 'CPlus Result'

        sanityWorksheet.column_dimensions['E'].width = 16
        sanityWorksheet['E1'].font = celBold
        sanityWorksheet['E1'].alignment = celCent
        sanityWorksheet['E1'] = 'WebSOP Result'

        sanityWorksheet.column_dimensions['F'].width = 15
        sanityWorksheet['F1'].font = celBold
        sanityWorksheet['F1'].alignment = celCent
        sanityWorksheet['F1'] = 'IOM Data'

        sanityWorksheet.column_dimensions['G'].width = 18
        sanityWorksheet['G1'].font = celBold
        sanityWorksheet['G1'].alignment = celCent
        sanityWorksheet['G1'] = 'Destination System'

        sanityWorksheet.column_dimensions['H'].width = 14
        sanityWorksheet['H1'].font = celBold
        sanityWorksheet['H1'].alignment = celCent
        sanityWorksheet['H1'] = 'Order Status'

        sanityWorksheet.column_dimensions['I'].width = 14
        sanityWorksheet['I1'].font = celBold
        sanityWorksheet['I1'].alignment = celCent
        sanityWorksheet['I1'] = 'OBAN Result'

        sanityWorksheet.column_dimensions['J'].width = 14
        sanityWorksheet['J1'].font = celBold
        sanityWorksheet['J1'].alignment = celCent
        sanityWorksheet['J1'] = 'Time Stamp'

        sanityWorkbook.save(currDir+"/Results/"+testDat+"/"+testEnv+'/'+testEnv+"_SR.xlsx")
        sanityWorkbook.close()


main()
