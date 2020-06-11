"""
"""

import openpyxl
import xlrd
import os
from os import path


def main():

    currDir = os.getcwd()
    srce = currDir + ("/Output/outputvars.xlsx")
    wBook1 = xlrd.open_workbook(srce)
    wSheet1 = wBook1.sheet_by_index(0)
    testEnv = wSheet1.cell_value(0,3)
    testTim = wSheet1.cell_value(0,1)
    
    srceOrdTyp = wSheet.cell_value(1,0)
    srceOrdNum = wSheet.cell_value(1,1)
    srceTN = wSheet.cell_value(1,2)
    srceCPlus = wSheet.cell_value(1,3)
    srceWebSOP = wSheet.cell_value(1,4)
    srceIOM = wSheet.cell_value(1,5)
    srceDSyst = wSheet.cell_value(1,6)
    srceOStat = wSheet.cell_value(1,7)
    srceOBAN = wSheet.cell_value(1,8)
    srceTime = wSheet.cell_value(1,9)

    dest = currDir + "/" + testEnv + "_SR.xlsx"
    wBook2 = xlrd.open_workbook(dest)
    wSheet2 = wBook2.sheet_by_name(str(testTim))
    
    destOrdTyp = wSheet2.append
    destOrdNum = wSheet.cell_value(1,1)
    destTN = wSheet.cell_value(1,2)
    destCPlus = wSheet.cell_value(1,3)
    destWebSOP = wSheet.cell_value(1,4)
    destIOM = wSheet.cell_value(1,5)
    destDSyst = wSheet.cell_value(1,6)
    destOStat = wSheet.cell_value(1,7)
    destOBAN = wSheet.cell_value(1,8)
    destTime = wSheet.cell_value(1,9)
    
    #wb1 = openpyxl.load_workbook(srce)
    #ws1 = wb1.worksheets(str(testTim))

    #wb2 = openpyxl.load_workbook(dest)
    #ws2 = wb2.active

    #mc = ws1.max_column

    #for i in range (1, 1):
        #for j in range (1, mc + 1):
            #c = ws1.cell(row = 2, column = j)
            #ws2.append(row = i, column = j).value = c.value

    #wb2.save(str(dest))
