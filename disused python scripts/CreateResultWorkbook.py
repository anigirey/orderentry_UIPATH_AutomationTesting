"""
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import xlrd
import os
from os import path


def main():

    currDir = os.getcwd()
    loc = currDir+("/Output/outputvars.xlsx")
    wBook = xlrd.open_workbook(loc)
    wSheet = wBook.sheet_by_index(0)
    testEnv = str(wSheet.cell_value(0,3))
    testTim = str(wSheet.cell_value(0,1))
    testDat = str(wSheet.cell_value(0,0))
    matchedSheet = int()

    if not path.exists(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx'):
        resultWorkbook = Workbook()
        resultWorksheet = resultWorkbook.create_sheet(testTim)
        dust = resultWorkbook.get_sheet_by_name('Sheet')
        resultWorkbook.remove_sheet(dust)

        celBold = Font(size=13,bold=True,underline='single')
        celBoldL = Font(size=11,bold=True)
        celCent = Alignment(horizontal='center')
        celLeft = Alignment(horizontal='left')
        
        resultWorksheet.column_dimensions['A'].width=47
        resultWorksheet['A1'].font=celBold
        resultWorksheet['A1'].alignment=celCent
        resultWorksheet['A1']='Step Detail'
        
        resultWorksheet['A2'].font=celBoldL
        resultWorksheet['A2'].alignment=celLeft
        resultWorksheet['A2']='Sign On Page opened'

        resultWorksheet['A3'].font=celBoldL
        resultWorksheet['A3'].alignment=celLeft
        resultWorksheet['A3']='Home Page opened'

        resultWorksheet['A4'].font=celBoldL
        resultWorksheet['A4'].alignment=celLeft
        resultWorksheet['A4']='Customer Information Page opened'

        resultWorksheet['A5'].font=celBoldL
        resultWorksheet['A5'].alignment=celLeft
        resultWorksheet['A5']='Win Back/Win Over Page opened'

        resultWorksheet['A6'].font=celBoldL
        resultWorksheet['A6'].alignment=celLeft
        resultWorksheet['A6']='Service Address Validation Page opened'

        resultWorksheet['A7'].font=celBoldL
        resultWorksheet['A7'].alignment=celLeft
        resultWorksheet['A7']='Facility Check and Results Page opened'

        resultWorksheet['A8'].font=celBoldL
        resultWorksheet['A8'].alignment=celLeft
        resultWorksheet['A8']='Primary Listing Page opened'

        resultWorksheet['A9'].font=celBoldL
        resultWorksheet['A9'].alignment=celLeft
        resultWorksheet['A9']='Product Pricing Page opened'

        resultWorksheet['A10'].font=celBoldL
        resultWorksheet['A10'].alignment=celLeft
        resultWorksheet['A10']='Billing Information Page opened'

        resultWorksheet['A11'].font=celBoldL
        resultWorksheet['A11'].alignment=celLeft
        resultWorksheet['A11']='Business Credit Application Page opened'

        resultWorksheet['A12'].font=celBoldL
        resultWorksheet['A12'].alignment=celLeft
        resultWorksheet['A12']='Credit Information Page opened'

        resultWorksheet['A13'].font=celBoldL
        resultWorksheet['A13'].alignment=celLeft
        resultWorksheet['A13']='Credit Decision Page opened'

        resultWorksheet['A14'].font=celBoldL
        resultWorksheet['A14'].alignment=celLeft
        resultWorksheet['A14']='Service and Equipment Page opened'

        resultWorksheet['A15'].font=celBoldL
        resultWorksheet['A15'].alignment=celLeft
        resultWorksheet['A15']='Product Summary Page opened'

        resultWorksheet['A16'].font=celBoldL
        resultWorksheet['A16'].alignment=celLeft
        resultWorksheet['A16']='Configure Product Page opened'

        resultWorksheet['A17'].font=celBoldL
        resultWorksheet['A17'].alignment=celLeft
        resultWorksheet['A17']='Configure Order Page opened'

        resultWorksheet['A18'].font=celBoldL
        resultWorksheet['A18'].alignment=celLeft
        resultWorksheet['A18']='Configure OLFIDs Page opened'

        resultWorksheet['A19'].font=celBoldL
        resultWorksheet['A19'].alignment=celLeft
        resultWorksheet['A19']='Appointment Scheduler Page opened'

        resultWorksheet['A20'].font=celBoldL
        resultWorksheet['A20'].alignment=celLeft
        resultWorksheet['A20']='Deposit/Advance Payment Page opened'

        resultWorksheet['A21'].font=celBoldL
        resultWorksheet['A21'].alignment=celLeft
        resultWorksheet['A21']='Deposit/Advance Payment Information Page opened'

        resultWorksheet['A22'].font=celBoldL
        resultWorksheet['A22'].alignment=celLeft
        resultWorksheet['A22']='Deposit/Advance Payment Success Page opened'

        resultWorksheet['A23'].font=celBoldL
        resultWorksheet['A23'].alignment=celLeft
        resultWorksheet['A23']='Order Detail Page opened'

        resultWorksheet['A24'].font=celBoldL
        resultWorksheet['A24'].alignment=celLeft
        resultWorksheet['A24']='Order Validation Confirmation'

        resultWorksheet['A26'].font=celBoldL
        resultWorksheet['A26'].alignment=celLeft
        resultWorksheet['A26']='Reason for Failure'

        # 1FR init
        resultWorksheet.column_dimensions['B'].width=15
        resultWorksheet['B1'].font=celBold
        resultWorksheet['B1'].alignment=celCent
        resultWorksheet['B1']='1FR'
        resultWorksheet['B2']='Not Tested'
        resultWorksheet['B3']='Not Tested'
        resultWorksheet['B4']='Not Tested'
        resultWorksheet['B5']='Not Tested'
        resultWorksheet['B6']='Not Tested'
        resultWorksheet['B7']='Not Tested'
        resultWorksheet['B8']='Not Tested'
        resultWorksheet['B9']='Not Tested'
        resultWorksheet['B10']='Not Tested'
        resultWorksheet['B11']='Not Tested'
        resultWorksheet['B12']='Not Tested'
        resultWorksheet['B13']='Not Tested'
        resultWorksheet['B14']='Not Tested'
        resultWorksheet['B15']='Not Tested'
        resultWorksheet['B16']='Not Tested'
        resultWorksheet['B17']='Not Tested'
        resultWorksheet['B18']='Not Tested'
        resultWorksheet['B19']='Not Tested'
        resultWorksheet['B20']='Not Tested'
        resultWorksheet['B21']='Not Tested'
        resultWorksheet['B22']='Not Tested'
        resultWorksheet['B23']='Not Tested'
        resultWorksheet['B24']='Not Tested'
        resultWorksheet['B26']='---'

        # 1FB init
        resultWorksheet.column_dimensions['C'].width=15
        resultWorksheet['C1'].font=celBold
        resultWorksheet['C1'].alignment=celCent
        resultWorksheet['C1']='1FB'
        resultWorksheet['C2']='Not Tested'
        resultWorksheet['C3']='Not Tested'
        resultWorksheet['C4']='Not Tested'
        resultWorksheet['C5']='Not Tested'
        resultWorksheet['C6']='Not Tested'
        resultWorksheet['C7']='Not Tested'
        resultWorksheet['C8']='Not Tested'
        resultWorksheet['C9']='Not Tested'
        resultWorksheet['C10']='Not Tested'
        resultWorksheet['C11']='Not Tested'
        resultWorksheet['C12']='Not Tested'
        resultWorksheet['C13']='Not Tested'
        resultWorksheet['C14']='Not Tested'
        resultWorksheet['C15']='Not Tested'
        resultWorksheet['C16']='Not Tested'
        resultWorksheet['C17']='Not Tested'
        resultWorksheet['C18']='Not Tested'
        resultWorksheet['C19']='Not Tested'
        resultWorksheet['C20']='Not Tested'
        resultWorksheet['C21']='Not Tested'
        resultWorksheet['C22']='Not Tested'
        resultWorksheet['C23']='Not Tested'
        resultWorksheet['C24']='Not Tested'
        resultWorksheet['C26']='---'

        # HSI init
        resultWorksheet.column_dimensions['D'].width=15
        resultWorksheet['D1'].font=celBold
        resultWorksheet['D1'].alignment=celCent
        resultWorksheet['D1']='HSI'
        resultWorksheet['D2']='Not Tested'
        resultWorksheet['D3']='Not Tested'
        resultWorksheet['D4']='Not Tested'
        resultWorksheet['D5']='Not Tested'
        resultWorksheet['D6']='Not Tested'
        resultWorksheet['D7']='Not Tested'
        resultWorksheet['D8']='Not Tested'
        resultWorksheet['D9']='Not Tested'
        resultWorksheet['D10']='Not Tested'
        resultWorksheet['D11']='Not Tested'
        resultWorksheet['D12']='Not Tested'
        resultWorksheet['D13']='Not Tested'
        resultWorksheet['D14']='Not Tested'
        resultWorksheet['D15']='Not Tested'
        resultWorksheet['D16']='Not Tested'
        resultWorksheet['D17']='Not Tested'
        resultWorksheet['D18']='Not Tested'
        resultWorksheet['D19']='Not Tested'
        resultWorksheet['D20']='Not Tested'
        resultWorksheet['D21']='Not Tested'
        resultWorksheet['D22']='Not Tested'
        resultWorksheet['D23']='Not Tested'
        resultWorksheet['D24']='Not Tested'
        resultWorksheet['D26']='---'

        # PRISM init
        resultWorksheet.column_dimensions['E'].width=15
        resultWorksheet['E1'].font=celBold
        resultWorksheet['E1'].alignment=celCent
        resultWorksheet['E1']='PRISM'
        resultWorksheet['E2']='Not Tested'
        resultWorksheet['E3']='Not Tested'
        resultWorksheet['E4']='Not Tested'
        resultWorksheet['E5']='Not Tested'
        resultWorksheet['E6']='Not Tested'
        resultWorksheet['E7']='Not Tested'
        resultWorksheet['E8']='Not Tested'
        resultWorksheet['E9']='Not Tested'
        resultWorksheet['E10']='Not Tested'
        resultWorksheet['E11']='Not Tested'
        resultWorksheet['E12']='Not Tested'
        resultWorksheet['E13']='Not Tested'
        resultWorksheet['E14']='Not Tested'
        resultWorksheet['E15']='Not Tested'
        resultWorksheet['E16']='Not Tested'
        resultWorksheet['E17']='Not Tested'
        resultWorksheet['E18']='Not Tested'
        resultWorksheet['E19']='Not Tested'
        resultWorksheet['E20']='Not Tested'
        resultWorksheet['E21']='Not Tested'
        resultWorksheet['E22']='Not Tested'
        resultWorksheet['E23']='Not Tested'
        resultWorksheet['E24']='Not Tested'
        resultWorksheet['E26']='---'

        resultWorkbook.save(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx')
        resultWorkbook.close()

    else:
        
        resultWorkbook = openpyxl.load_workbook(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx')
                
        for i in resultWorkbook.sheetnames:
            if (str(i) == testTim):
                matchedSheet = int(1)
                break
            else:
                matchedSheet = int(0)

        if (matchedSheet == int(0)):

            resultWorksheet = resultWorkbook.create_sheet(testTim)
            
            celBold = Font(size=13,bold=True,underline='single')
            celBoldL = Font(size=11,bold=True)
            celCent = Alignment(horizontal='center')
            celLeft = Alignment(horizontal='left')
            
            resultWorksheet.column_dimensions['A'].width=47
            resultWorksheet['A1'].font=celBold
            resultWorksheet['A1'].alignment=celCent
            resultWorksheet['A1']='Step Detail'
            
            resultWorksheet['A2'].font=celBoldL
            resultWorksheet['A2'].alignment=celLeft
            resultWorksheet['A2']='Sign On Page opened'

            resultWorksheet['A3'].font=celBoldL
            resultWorksheet['A3'].alignment=celLeft
            resultWorksheet['A3']='Home Page opened'

            resultWorksheet['A4'].font=celBoldL
            resultWorksheet['A4'].alignment=celLeft
            resultWorksheet['A4']='Customer Information Page opened'

            resultWorksheet['A5'].font=celBoldL
            resultWorksheet['A5'].alignment=celLeft
            resultWorksheet['A5']='Win Back/Win Over Page opened'

            resultWorksheet['A6'].font=celBoldL
            resultWorksheet['A6'].alignment=celLeft
            resultWorksheet['A6']='Service Address Validation Page opened'

            resultWorksheet['A7'].font=celBoldL
            resultWorksheet['A7'].alignment=celLeft
            resultWorksheet['A7']='Facility Check and Results Page opened'

            resultWorksheet['A8'].font=celBoldL
            resultWorksheet['A8'].alignment=celLeft
            resultWorksheet['A8']='Primary Listing Page opened'

            resultWorksheet['A9'].font=celBoldL
            resultWorksheet['A9'].alignment=celLeft
            resultWorksheet['A9']='Product Pricing Page opened'

            resultWorksheet['A10'].font=celBoldL
            resultWorksheet['A10'].alignment=celLeft
            resultWorksheet['A10']='Billing Information Page opened'

            resultWorksheet['A11'].font=celBoldL
            resultWorksheet['A11'].alignment=celLeft
            resultWorksheet['A11']='Business Credit Application Page opened'

            resultWorksheet['A12'].font=celBoldL
            resultWorksheet['A12'].alignment=celLeft
            resultWorksheet['A12']='Credit Information Page opened'

            resultWorksheet['A13'].font=celBoldL
            resultWorksheet['A13'].alignment=celLeft
            resultWorksheet['A13']='Credit Decision Page opened'

            resultWorksheet['A14'].font=celBoldL
            resultWorksheet['A14'].alignment=celLeft
            resultWorksheet['A14']='Service and Equipment Page opened'

            resultWorksheet['A15'].font=celBoldL
            resultWorksheet['A15'].alignment=celLeft
            resultWorksheet['A15']='Product Summary Page opened'

            resultWorksheet['A16'].font=celBoldL
            resultWorksheet['A16'].alignment=celLeft
            resultWorksheet['A16']='Configure Product Page opened'

            resultWorksheet['A17'].font=celBoldL
            resultWorksheet['A17'].alignment=celLeft
            resultWorksheet['A17']='Configure Order Page opened'

            resultWorksheet['A18'].font=celBoldL
            resultWorksheet['A18'].alignment=celLeft
            resultWorksheet['A18']='Configure OLFIDs Page opened'

            resultWorksheet['A19'].font=celBoldL
            resultWorksheet['A19'].alignment=celLeft
            resultWorksheet['A19']='Appointment Scheduler Page opened'

            resultWorksheet['A20'].font=celBoldL
            resultWorksheet['A20'].alignment=celLeft
            resultWorksheet['A20']='Deposit/Advance Payment Page opened'

            resultWorksheet['A21'].font=celBoldL
            resultWorksheet['A21'].alignment=celLeft
            resultWorksheet['A21']='Deposit/Advance Payment Information Page opened'

            resultWorksheet['A22'].font=celBoldL
            resultWorksheet['A22'].alignment=celLeft
            resultWorksheet['A22']='Deposit/Advance Payment Success Page opened'

            resultWorksheet['A23'].font=celBoldL
            resultWorksheet['A23'].alignment=celLeft
            resultWorksheet['A23']='Order Detail Page opened'

            resultWorksheet['A24'].font=celBoldL
            resultWorksheet['A24'].alignment=celLeft
            resultWorksheet['A24']='Order Validation Confirmation'

            resultWorksheet['A26'].font=celBoldL
            resultWorksheet['A26'].alignment=celLeft
            resultWorksheet['A26']='Reason for Failure'

            # 1FR init
            resultWorksheet.column_dimensions['B'].width=15
            resultWorksheet['B1'].font=celBold
            resultWorksheet['B1'].alignment=celCent
            resultWorksheet['B1']='1FR'
            resultWorksheet['B2']='Not Tested'
            resultWorksheet['B3']='Not Tested'
            resultWorksheet['B4']='Not Tested'
            resultWorksheet['B5']='Not Tested'
            resultWorksheet['B6']='Not Tested'
            resultWorksheet['B7']='Not Tested'
            resultWorksheet['B8']='Not Tested'
            resultWorksheet['B9']='Not Tested'
            resultWorksheet['B10']='Not Tested'
            resultWorksheet['B11']='Not Tested'
            resultWorksheet['B12']='Not Tested'
            resultWorksheet['B13']='Not Tested'
            resultWorksheet['B14']='Not Tested'
            resultWorksheet['B15']='Not Tested'
            resultWorksheet['B16']='Not Tested'
            resultWorksheet['B17']='Not Tested'
            resultWorksheet['B18']='Not Tested'
            resultWorksheet['B19']='Not Tested'
            resultWorksheet['B20']='Not Tested'
            resultWorksheet['B21']='Not Tested'
            resultWorksheet['B22']='Not Tested'
            resultWorksheet['B23']='Not Tested'
            resultWorksheet['B24']='Not Tested'
            resultWorksheet['B26']='---'

            # 1FB init
            resultWorksheet.column_dimensions['C'].width=15
            resultWorksheet['C1'].font=celBold
            resultWorksheet['C1'].alignment=celCent
            resultWorksheet['C1']='1FB'
            resultWorksheet['C2']='Not Tested'
            resultWorksheet['C3']='Not Tested'
            resultWorksheet['C4']='Not Tested'
            resultWorksheet['C5']='Not Tested'
            resultWorksheet['C6']='Not Tested'
            resultWorksheet['C7']='Not Tested'
            resultWorksheet['C8']='Not Tested'
            resultWorksheet['C9']='Not Tested'
            resultWorksheet['C10']='Not Tested'
            resultWorksheet['C11']='Not Tested'
            resultWorksheet['C12']='Not Tested'
            resultWorksheet['C13']='Not Tested'
            resultWorksheet['C14']='Not Tested'
            resultWorksheet['C15']='Not Tested'
            resultWorksheet['C16']='Not Tested'
            resultWorksheet['C17']='Not Tested'
            resultWorksheet['C18']='Not Tested'
            resultWorksheet['C19']='Not Tested'
            resultWorksheet['C20']='Not Tested'
            resultWorksheet['C21']='Not Tested'
            resultWorksheet['C22']='Not Tested'
            resultWorksheet['C23']='Not Tested'
            resultWorksheet['C24']='Not Tested'
            resultWorksheet['C26']='---'

            # HSI init
            resultWorksheet.column_dimensions['D'].width=15
            resultWorksheet['D1'].font=celBold
            resultWorksheet['D1'].alignment=celCent
            resultWorksheet['D1']='HSI'
            resultWorksheet['D2']='Not Tested'
            resultWorksheet['D3']='Not Tested'
            resultWorksheet['D4']='Not Tested'
            resultWorksheet['D5']='Not Tested'
            resultWorksheet['D6']='Not Tested'
            resultWorksheet['D7']='Not Tested'
            resultWorksheet['D8']='Not Tested'
            resultWorksheet['D9']='Not Tested'
            resultWorksheet['D10']='Not Tested'
            resultWorksheet['D11']='Not Tested'
            resultWorksheet['D12']='Not Tested'
            resultWorksheet['D13']='Not Tested'
            resultWorksheet['D14']='Not Tested'
            resultWorksheet['D15']='Not Tested'
            resultWorksheet['D16']='Not Tested'
            resultWorksheet['D17']='Not Tested'
            resultWorksheet['D18']='Not Tested'
            resultWorksheet['D19']='Not Tested'
            resultWorksheet['D20']='Not Tested'
            resultWorksheet['D21']='Not Tested'
            resultWorksheet['D22']='Not Tested'
            resultWorksheet['D23']='Not Tested'
            resultWorksheet['D24']='Not Tested'
            resultWorksheet['D26']='---'

            # PRISM init
            resultWorksheet.column_dimensions['E'].width=15
            resultWorksheet['E1'].font=celBold
            resultWorksheet['E1'].alignment=celCent
            resultWorksheet['E1']='PRISM'
            resultWorksheet['E2']='Not Tested'
            resultWorksheet['E3']='Not Tested'
            resultWorksheet['E4']='Not Tested'
            resultWorksheet['E5']='Not Tested'
            resultWorksheet['E6']='Not Tested'
            resultWorksheet['E7']='Not Tested'
            resultWorksheet['E8']='Not Tested'
            resultWorksheet['E9']='Not Tested'
            resultWorksheet['E10']='Not Tested'
            resultWorksheet['E11']='Not Tested'
            resultWorksheet['E12']='Not Tested'
            resultWorksheet['E13']='Not Tested'
            resultWorksheet['E14']='Not Tested'
            resultWorksheet['E15']='Not Tested'
            resultWorksheet['E16']='Not Tested'
            resultWorksheet['E17']='Not Tested'
            resultWorksheet['E18']='Not Tested'
            resultWorksheet['E19']='Not Tested'
            resultWorksheet['E20']='Not Tested'
            resultWorksheet['E21']='Not Tested'
            resultWorksheet['E22']='Not Tested'
            resultWorksheet['E23']='Not Tested'
            resultWorksheet['E24']='Not Tested'
            resultWorksheet['E26']='---'

            resultWorkbook.save(currDir+'/Results/'+testDat+'/'+testEnv+'/Summary File.xlsx')
            resultWorkbook.close()


main()
